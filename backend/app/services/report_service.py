from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime
from sqlalchemy.orm import Session
from app.models import Tenant, Payment, Invoice, Unit, TenantStatus, PaymentStatus


# ── COLORS ──
NAVY   = colors.HexColor('#0f172a')
AMBER  = colors.HexColor('#d97706')
SLATE  = colors.HexColor('#64748b')
LIGHT  = colors.HexColor('#f8fafc')
WHITE  = colors.white
BORDER = colors.HexColor('#e2e8f0')


def fmt(n):
    return f"KES {float(n or 0):,.0f}"


def build_pdf_header(elements, styles, title, subtitle):
    elements.append(Spacer(1, 6*mm))
    elements.append(Paragraph(
        "ABISAKI'S RENTAL MANAGEMENT SYSTEM",
        ParagraphStyle('brand', fontSize=9, textColor=SLATE, alignment=TA_CENTER,
                       fontName='Helvetica', letterSpacing=1.5)
    ))
    elements.append(Spacer(1, 3*mm))
    elements.append(Paragraph(
        title,
        ParagraphStyle('title', fontSize=18, textColor=NAVY, alignment=TA_CENTER,
                       fontName='Helvetica-Bold', spaceAfter=2*mm)
    ))
    elements.append(Paragraph(
        subtitle,
        ParagraphStyle('sub', fontSize=9, textColor=SLATE, alignment=TA_CENTER,
                       fontName='Helvetica')
    ))
    elements.append(Spacer(1, 6*mm))

    # divider
    elements.append(Table(
        [['']],
        colWidths=[170*mm],
        style=TableStyle([
            ('LINEBELOW', (0,0), (-1,-1), 2, AMBER),
        ])
    ))
    elements.append(Spacer(1, 6*mm))


# ────────────────────────────────────────────
# 1. PER-TENANT STATEMENT PDF
# ────────────────────────────────────────────
def generate_tenant_statement_pdf(tenant_id: int, db: Session) -> BytesIO:
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise ValueError("Tenant not found")

    unit   = db.query(Unit).filter(Unit.id == tenant.unit_id).first()
    payments = db.query(Payment).filter(
        Payment.tenant_id == tenant_id,
        Payment.status == PaymentStatus.confirmed
    ).order_by(Payment.timestamp.asc()).all()
    invoices = db.query(Invoice).filter(
        Invoice.tenant_id == tenant_id
    ).order_by(Invoice.created_at.asc()).all()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=15*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    elements = []

    build_pdf_header(
        elements, styles,
        "Tenant Account Statement",
        f"Generated on {datetime.now().strftime('%d %B %Y at %H:%M')}"
    )

    # tenant info box
    info_data = [
        ['Tenant Name', tenant.name, 'Unit', unit.unit_number if unit else '—'],
        ['Phone', tenant.phone_number, 'Monthly Rent', fmt(unit.rent_amount) if unit else '—'],
        ['Lease Start', str(tenant.lease_start_date), 'Lease End', str(tenant.lease_end_date)],
        ['Current Balance', fmt(tenant.balance), 'Account Status', 'Active' if tenant.balance <= 0 else 'Outstanding'],
    ]
    info_table = Table(info_data, colWidths=[40*mm, 60*mm, 35*mm, 45*mm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (0,-1), LIGHT),
        ('BACKGROUND',   (2,0), (2,-1), LIGHT),
        ('FONTNAME',     (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME',     (2,0), (2,-1), 'Helvetica-Bold'),
        ('FONTNAME',     (1,0), (1,-1), 'Helvetica'),
        ('FONTNAME',     (3,0), (3,-1), 'Helvetica'),
        ('FONTSIZE',     (0,0), (-1,-1), 9),
        ('TEXTCOLOR',    (0,0), (0,-1), NAVY),
        ('TEXTCOLOR',    (2,0), (2,-1), NAVY),
        ('TEXTCOLOR',    (1,0), (1,-1), SLATE),
        ('TEXTCOLOR',    (3,0), (3,-1), SLATE),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [WHITE, LIGHT]),
        ('BOX',          (0,0), (-1,-1), 0.5, BORDER),
        ('INNERGRID',    (0,0), (-1,-1), 0.3, BORDER),
        ('PADDING',      (0,0), (-1,-1), 6),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 8*mm))

    # charges section
    elements.append(Paragraph(
        "BILLING CHARGES",
        ParagraphStyle('sh', fontSize=8, textColor=AMBER, fontName='Helvetica-Bold',
                       letterSpacing=1, spaceAfter=3*mm)
    ))
    if invoices:
        inv_data = [['#', 'Period', 'Amount Charged', 'Status', 'Date']]
        for inv in invoices:
            inv_data.append([
                str(inv.id),
                inv.billing_period,
                fmt(inv.amount),
                inv.status.value if hasattr(inv.status,'value') else str(inv.status),
                inv.created_at.strftime('%d %b %Y') if inv.created_at else '—'
            ])
        inv_table = Table(inv_data, colWidths=[15*mm, 45*mm, 40*mm, 35*mm, 35*mm])
        inv_table.setStyle(TableStyle([
            ('BACKGROUND',   (0,0), (-1,0), NAVY),
            ('TEXTCOLOR',    (0,0), (-1,0), WHITE),
            ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0,0), (-1,-1), 8.5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [WHITE, LIGHT]),
            ('TEXTCOLOR',    (0,1), (-1,-1), SLATE),
            ('FONTNAME',     (0,1), (-1,-1), 'Helvetica'),
            ('ALIGN',        (2,0), (2,-1), 'RIGHT'),
            ('BOX',          (0,0), (-1,-1), 0.5, BORDER),
            ('INNERGRID',    (0,0), (-1,-1), 0.3, BORDER),
            ('PADDING',      (0,0), (-1,-1), 6),
            ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ]))
        elements.append(inv_table)
    else:
        elements.append(Paragraph("No billing charges recorded.", 
                        ParagraphStyle('note', fontSize=9, textColor=SLATE)))
    elements.append(Spacer(1, 8*mm))

    # payments section
    elements.append(Paragraph(
        "PAYMENT RECEIPTS",
        ParagraphStyle('sh', fontSize=8, textColor=AMBER, fontName='Helvetica-Bold',
                       letterSpacing=1, spaceAfter=3*mm)
    ))
    if payments:
        total_paid = sum(p.amount for p in payments)
        pay_data = [['Receipt No.', 'Method', 'Amount Paid', 'Phone', 'Date']]
        for p in payments:
            pay_data.append([
                p.mpesa_receipt_number or '—',
                p.payment_method.value if hasattr(p.payment_method,'value') else str(p.payment_method),
                fmt(p.amount),
                p.phone_number or '—',
                p.timestamp.strftime('%d %b %Y %H:%M') if p.timestamp else '—'
            ])
        pay_data.append(['', '', fmt(total_paid), '', ''])
        pay_table = Table(pay_data, colWidths=[35*mm, 35*mm, 40*mm, 35*mm, 35*mm])
        pay_table.setStyle(TableStyle([
            ('BACKGROUND',   (0,0),  (-1,0),  NAVY),
            ('TEXTCOLOR',    (0,0),  (-1,0),  WHITE),
            ('FONTNAME',     (0,0),  (-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',     (0,0),  (-1,-1), 8.5),
            ('ROWBACKGROUNDS',(0,1), (-1,-2), [WHITE, LIGHT]),
            ('TEXTCOLOR',    (0,1),  (-1,-2), SLATE),
            ('FONTNAME',     (0,1),  (-1,-1), 'Helvetica'),
            ('BACKGROUND',   (0,-1), (-1,-1), LIGHT),
            ('FONTNAME',     (2,-1), (2,-1),  'Helvetica-Bold'),
            ('TEXTCOLOR',    (2,-1), (2,-1),  NAVY),
            ('ALIGN',        (2,0),  (2,-1),  'RIGHT'),
            ('BOX',          (0,0),  (-1,-1), 0.5, BORDER),
            ('INNERGRID',    (0,0),  (-1,-1), 0.3, BORDER),
            ('PADDING',      (0,0),  (-1,-1), 6),
            ('VALIGN',       (0,0),  (-1,-1), 'MIDDLE'),
        ]))
        elements.append(pay_table)
    else:
        elements.append(Paragraph("No payments recorded.",
                        ParagraphStyle('note', fontSize=9, textColor=SLATE)))

    elements.append(Spacer(1, 8*mm))

    # summary box
    summary_data = [
        ['Total Charged', fmt(sum(i.amount for i in invoices))],
        ['Total Paid',    fmt(sum(p.amount for p in payments))],
        ['Balance Due',   fmt(tenant.balance)],
    ]
    summary_table = Table(summary_data, colWidths=[80*mm, 90*mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND',  (0,0),  (0,-1), LIGHT),
        ('FONTNAME',    (0,0),  (0,-1), 'Helvetica-Bold'),
        ('FONTNAME',    (1,0),  (1,-2), 'Helvetica'),
        ('FONTNAME',    (1,-1), (1,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR',   (0,0),  (0,-1), NAVY),
        ('TEXTCOLOR',   (1,-1), (1,-1), colors.HexColor('#e11d48')),
        ('FONTSIZE',    (0,0),  (-1,-1), 9.5),
        ('ALIGN',       (1,0),  (1,-1), 'RIGHT'),
        ('BOX',         (0,0),  (-1,-1), 1, AMBER),
        ('INNERGRID',   (0,0),  (-1,-1), 0.3, BORDER),
        ('PADDING',     (0,0),  (-1,-1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 10*mm))
    elements.append(Paragraph(
        "This is a computer-generated statement. No signature required.",
        ParagraphStyle('footer', fontSize=7.5, textColor=SLATE, alignment=TA_CENTER)
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ────────────────────────────────────────────
# 2. MONTHLY COLLECTIONS PDF
# ────────────────────────────────────────────
def generate_monthly_report_pdf(month: str, db: Session) -> BytesIO:
    # month format: "2026-06"
    try:
        dt = datetime.strptime(month, "%Y-%m")
        month_label = dt.strftime("%B %Y")
    except:
        month_label = month

    payments = db.query(Payment).filter(
        Payment.status == PaymentStatus.confirmed
    ).all()
    month_payments = [
        p for p in payments
        if p.timestamp and p.timestamp.strftime("%Y-%m") == month
    ]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=15*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    elements = []

    build_pdf_header(
        elements, styles,
        f"Monthly Collections Report",
        f"{month_label}  •  Generated {datetime.now().strftime('%d %B %Y')}"
    )

    if not month_payments:
        elements.append(Paragraph(
            f"No payments recorded for {month_label}.",
            ParagraphStyle('note', fontSize=10, textColor=SLATE, alignment=TA_CENTER)
        ))
        doc.build(elements)
        buffer.seek(0)
        return buffer

    total = sum(p.amount for p in month_payments)
    mpesa = sum(p.amount for p in month_payments if 'mpesa' in str(p.payment_method).lower() or 'stk' in str(p.payment_method).lower())
    cash  = sum(p.amount for p in month_payments if 'cash' in str(p.payment_method).lower() or 'manual' in str(p.payment_method).lower())

    # summary cards row
    summary_data = [
        ['Total Collected', 'M-Pesa Payments', 'Cash Payments', 'Transactions'],
        [fmt(total), fmt(mpesa), fmt(cash), str(len(month_payments))]
    ]
    summary_table = Table(summary_data, colWidths=[42*mm, 42*mm, 42*mm, 42*mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), NAVY),
        ('TEXTCOLOR',   (0,0), (-1,0), WHITE),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 8),
        ('BACKGROUND',  (0,1), (-1,1), LIGHT),
        ('TEXTCOLOR',   (0,1), (-1,1), NAVY),
        ('FONTNAME',    (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,1), (-1,1), 11),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('BOX',         (0,0), (-1,-1), 0.5, BORDER),
        ('INNERGRID',   (0,0), (-1,-1), 0.3, BORDER),
        ('PADDING',     (0,0), (-1,-1), 8),
        ('LINEBELOW',   (0,0), (-1,0), 2, AMBER),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 8*mm))

    elements.append(Paragraph(
        "TRANSACTION BREAKDOWN",
        ParagraphStyle('sh', fontSize=8, textColor=AMBER, fontName='Helvetica-Bold',
                       letterSpacing=1, spaceAfter=3*mm)
    ))

    pay_data = [['Receipt No.', 'Tenant ID', 'Amount', 'Method', 'Phone', 'Date']]
    for p in sorted(month_payments, key=lambda x: x.timestamp):
        pay_data.append([
            p.mpesa_receipt_number or '—',
            f"#{p.tenant_id}",
            fmt(p.amount),
            p.payment_method.value if hasattr(p.payment_method,'value') else str(p.payment_method),
            p.phone_number or '—',
            p.timestamp.strftime('%d %b %H:%M') if p.timestamp else '—'
        ])
    pay_data.append(['TOTAL', '', fmt(total), '', '', ''])

    pay_table = Table(pay_data, colWidths=[32*mm, 22*mm, 35*mm, 32*mm, 32*mm, 25*mm])
    pay_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  NAVY),
        ('TEXTCOLOR',     (0,0),  (-1,0),  WHITE),
        ('FONTNAME',      (0,0),  (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0),  (-1,-1), 8),
        ('ROWBACKGROUNDS',(0,1),  (-1,-2), [WHITE, LIGHT]),
        ('TEXTCOLOR',     (0,1),  (-1,-2), SLATE),
        ('FONTNAME',      (0,1),  (-1,-1), 'Helvetica'),
        ('BACKGROUND',    (0,-1), (-1,-1), LIGHT),
        ('FONTNAME',      (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR',     (0,-1), (-1,-1), NAVY),
        ('ALIGN',         (2,0),  (2,-1),  'RIGHT'),
        ('BOX',           (0,0),  (-1,-1), 0.5, BORDER),
        ('INNERGRID',     (0,0),  (-1,-1), 0.3, BORDER),
        ('PADDING',       (0,0),  (-1,-1), 5),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
    ]))
    elements.append(pay_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ────────────────────────────────────────────
# 3. FULL PORTFOLIO REPORT PDF
# ────────────────────────────────────────────
def generate_portfolio_report_pdf(db: Session) -> BytesIO:
    tenants = db.query(Tenant).filter(Tenant.status == TenantStatus.active).all()
    units   = db.query(Unit).all()
    payments = db.query(Payment).filter(Payment.status == PaymentStatus.confirmed).all()

    total_debt        = sum(t.balance for t in tenants if t.balance > 0)
    total_collections = sum(p.amount for p in payments)
    occupied          = len(tenants)
    total_units       = len(units)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=15*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    elements = []

    build_pdf_header(
        elements, styles,
        "Full Portfolio Report",
        f"Generated {datetime.now().strftime('%d %B %Y at %H:%M')}"
    )

    # portfolio summary
    summary_data = [
        ['Total Units', 'Occupied', 'Outstanding Debt', 'Total Collections'],
        [str(total_units), str(occupied), fmt(total_debt), fmt(total_collections)]
    ]
    summary_table = Table(summary_data, colWidths=[42*mm, 42*mm, 42*mm, 42*mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), NAVY),
        ('TEXTCOLOR',   (0,0), (-1,0), WHITE),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 8),
        ('BACKGROUND',  (0,1), (-1,1), LIGHT),
        ('TEXTCOLOR',   (0,1), (-1,1), NAVY),
        ('FONTNAME',    (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,1), (-1,1), 11),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('BOX',         (0,0), (-1,-1), 0.5, BORDER),
        ('INNERGRID',   (0,0), (-1,-1), 0.3, BORDER),
        ('PADDING',     (0,0), (-1,-1), 8),
        ('LINEBELOW',   (0,0), (-1,0), 2, AMBER),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 8*mm))

    # tenant ledger
    elements.append(Paragraph(
        "ACTIVE TENANT LEDGER",
        ParagraphStyle('sh', fontSize=8, textColor=AMBER, fontName='Helvetica-Bold',
                       letterSpacing=1, spaceAfter=3*mm)
    ))

    if tenants:
        t_data = [['Name', 'Phone', 'Unit', 'Balance', 'Lease End', 'Status']]
        for t in tenants:
            unit = db.query(Unit).filter(Unit.id == t.unit_id).first()
            t_data.append([
                t.name,
                t.phone_number,
                unit.unit_number if unit else '—',
                fmt(t.balance),
                str(t.lease_end_date),
                'Cleared' if t.balance <= 0 else 'Outstanding'
            ])
        t_table = Table(t_data, colWidths=[42*mm, 32*mm, 20*mm, 30*mm, 26*mm, 28*mm])
        t_table.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),  (-1,0),  NAVY),
            ('TEXTCOLOR',     (0,0),  (-1,0),  WHITE),
            ('FONTNAME',      (0,0),  (-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0,0),  (-1,-1), 8),
            ('ROWBACKGROUNDS',(0,1),  (-1,-1), [WHITE, LIGHT]),
            ('TEXTCOLOR',     (0,1),  (-1,-1), SLATE),
            ('FONTNAME',      (0,1),  (-1,-1), 'Helvetica'),
            ('ALIGN',         (3,0),  (3,-1),  'RIGHT'),
            ('BOX',           (0,0),  (-1,-1), 0.5, BORDER),
            ('INNERGRID',     (0,0),  (-1,-1), 0.3, BORDER),
            ('PADDING',       (0,0),  (-1,-1), 5),
            ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ]))
        elements.append(t_table)
    elements.append(Spacer(1, 8*mm))

    # units section
    elements.append(Paragraph(
        "UNITS OVERVIEW",
        ParagraphStyle('sh', fontSize=8, textColor=AMBER, fontName='Helvetica-Bold',
                       letterSpacing=1, spaceAfter=3*mm)
    ))
    u_data = [['Unit No.', 'Monthly Rent', 'Status', 'Occupancy']]
    for u in units:
        is_occ = any(t.unit_id == u.id for t in tenants)
        u_data.append([
            u.unit_number,
            fmt(u.rent_amount),
            str(u.status.value) if hasattr(u.status,'value') else str(u.status),
            'Occupied' if is_occ else 'Vacant'
        ])
    u_table = Table(u_data, colWidths=[40*mm, 45*mm, 42*mm, 42*mm])
    u_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  NAVY),
        ('TEXTCOLOR',     (0,0),  (-1,0),  WHITE),
        ('FONTNAME',      (0,0),  (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0),  (-1,-1), 8),
        ('ROWBACKGROUNDS',(0,1),  (-1,-1), [WHITE, LIGHT]),
        ('TEXTCOLOR',     (0,1),  (-1,-1), SLATE),
        ('FONTNAME',      (0,1),  (-1,-1), 'Helvetica'),
        ('BOX',           (0,0),  (-1,-1), 0.5, BORDER),
        ('INNERGRID',     (0,0),  (-1,-1), 0.3, BORDER),
        ('PADDING',       (0,0),  (-1,-1), 5),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
    ]))
    elements.append(u_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ────────────────────────────────────────────
# EXCEL EXPORTS
# ────────────────────────────────────────────
EXCEL_NAVY  = "0F172A"
EXCEL_AMBER = "D97706"
EXCEL_LIGHT = "F8FAFC"
EXCEL_WHITE = "FFFFFF"
EXCEL_SLATE = "64748B"

def excel_header_style():
    return {
        'font':      Font(bold=True, color=EXCEL_WHITE, size=10),
        'fill':      PatternFill("solid", fgColor=EXCEL_NAVY),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border':    Border(
            bottom=Side(style='medium', color=EXCEL_AMBER)
        )
    }

def excel_row_style(row_idx):
    fill_color = EXCEL_LIGHT if row_idx % 2 == 0 else EXCEL_WHITE
    return {
        'fill':      PatternFill("solid", fgColor=fill_color),
        'font':      Font(size=9, color=EXCEL_SLATE),
        'alignment': Alignment(vertical='center'),
        'border':    Border(
            bottom=Side(style='thin', color='E2E8F0')
        )
    }

def apply_style(cell, style):
    for attr, val in style.items():
        setattr(cell, attr, val)

def generate_tenant_statement_excel(tenant_id: int, db: Session) -> BytesIO:
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise ValueError("Tenant not found")
    unit     = db.query(Unit).filter(Unit.id == tenant.unit_id).first()
    payments = db.query(Payment).filter(
        Payment.tenant_id == tenant_id,
        Payment.status == PaymentStatus.confirmed
    ).order_by(Payment.timestamp.asc()).all()
    invoices = db.query(Invoice).filter(
        Invoice.tenant_id == tenant_id
    ).order_by(Invoice.created_at.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tenant Statement"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18

    # title
    ws.merge_cells('A1:E1')
    ws['A1'] = "ABISAKI'S RENTAL MANAGEMENT SYSTEM"
    ws['A1'].font      = Font(bold=True, size=11, color=EXCEL_NAVY)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Tenant Account Statement — Generated {datetime.now().strftime('%d %B %Y')}"
    ws['A2'].font      = Font(size=9, color=EXCEL_SLATE)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 6

    # info
    info = [
        ('Tenant Name', tenant.name, 'Unit', unit.unit_number if unit else '—'),
        ('Phone', tenant.phone_number, 'Monthly Rent', f"KES {float(unit.rent_amount or 0):,.0f}" if unit else '—'),
        ('Lease Start', str(tenant.lease_start_date), 'Lease End', str(tenant.lease_end_date)),
        ('Balance Due', f"KES {float(tenant.balance or 0):,.0f}", 'Status', 'Cleared' if tenant.balance <= 0 else 'Outstanding'),
    ]
    for r, row in enumerate(info, start=4):
        ws.cell(r, 1, row[0]).font = Font(bold=True, size=9, color=EXCEL_NAVY)
        ws.cell(r, 2, row[1]).font = Font(size=9, color=EXCEL_SLATE)
        ws.cell(r, 3, row[2]).font = Font(bold=True, size=9, color=EXCEL_NAVY)
        ws.cell(r, 4, row[3]).font = Font(size=9, color=EXCEL_SLATE)
        ws.row_dimensions[r].height = 16

    # charges
    row = 9
    ws.cell(row, 1, "BILLING CHARGES").font = Font(bold=True, size=9, color=EXCEL_AMBER)
    row += 1
    headers = ['#', 'Period', 'Amount', 'Status', 'Date']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        apply_style(cell, excel_header_style())
    ws.row_dimensions[row].height = 18
    row += 1
    for i, inv in enumerate(invoices):
        data = [
            str(inv.id), inv.billing_period,
            f"KES {float(inv.amount or 0):,.0f}",
            str(inv.status.value) if hasattr(inv.status,'value') else str(inv.status),
            inv.created_at.strftime('%d %b %Y') if inv.created_at else '—'
        ]
        style = excel_row_style(i)
        for c, v in enumerate(data, 1):
            cell = ws.cell(row, c, v)
            apply_style(cell, style)
        ws.row_dimensions[row].height = 15
        row += 1

    # payments
    row += 1
    ws.cell(row, 1, "PAYMENT RECEIPTS").font = Font(bold=True, size=9, color=EXCEL_AMBER)
    row += 1
    headers = ['Receipt No.', 'Method', 'Amount', 'Phone', 'Date']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        apply_style(cell, excel_header_style())
    ws.row_dimensions[row].height = 18
    row += 1
    for i, p in enumerate(payments):
        data = [
            p.mpesa_receipt_number or '—',
            p.payment_method.value if hasattr(p.payment_method,'value') else str(p.payment_method),
            f"KES {float(p.amount or 0):,.0f}",
            p.phone_number or '—',
            p.timestamp.strftime('%d %b %Y %H:%M') if p.timestamp else '—'
        ]
        style = excel_row_style(i)
        for c, v in enumerate(data, 1):
            cell = ws.cell(row, c, v)
            apply_style(cell, style)
        ws.row_dimensions[row].height = 15
        row += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_monthly_report_excel(month: str, db: Session) -> BytesIO:
    try:
        dt = datetime.strptime(month, "%Y-%m")
        month_label = dt.strftime("%B %Y")
    except:
        month_label = month

    payments = db.query(Payment).filter(
        Payment.status == PaymentStatus.confirmed
    ).all()
    month_payments = [
        p for p in payments
        if p.timestamp and p.timestamp.strftime("%Y-%m") == month
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Collections {month_label}"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Monthly Collections Report — {month_label}"
    ws['A1'].font      = Font(bold=True, size=12, color=EXCEL_NAVY)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Generated {datetime.now().strftime('%d %B %Y at %H:%M')}"
    ws['A2'].font      = Font(size=9, color=EXCEL_SLATE)
    ws['A2'].alignment = Alignment(horizontal='center')

    row = 4
    headers = ['Receipt No.', 'Tenant ID', 'Amount', 'Method', 'Phone', 'Date']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        apply_style(cell, excel_header_style())
    ws.row_dimensions[row].height = 18
    row += 1

    for i, p in enumerate(sorted(month_payments, key=lambda x: x.timestamp)):
        data = [
            p.mpesa_receipt_number or '—',
            f"#{p.tenant_id}",
            f"KES {float(p.amount or 0):,.0f}",
            p.payment_method.value if hasattr(p.payment_method,'value') else str(p.payment_method),
            p.phone_number or '—',
            p.timestamp.strftime('%d %b %Y %H:%M') if p.timestamp else '—'
        ]
        style = excel_row_style(i)
        for c, v in enumerate(data, 1):
            cell = ws.cell(row, c, v)
            apply_style(cell, style)
        ws.row_dimensions[row].height = 15
        row += 1

    # total row
    total = sum(p.amount for p in month_payments)
    ws.cell(row, 1, 'TOTAL').font      = Font(bold=True, size=9, color=EXCEL_NAVY)
    ws.cell(row, 3, f"KES {total:,.0f}").font = Font(bold=True, size=9, color=EXCEL_NAVY)
    ws.row_dimensions[row].height = 16

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_portfolio_report_excel(db: Session) -> BytesIO:
    tenants  = db.query(Tenant).filter(Tenant.status == TenantStatus.active).all()
    units    = db.query(Unit).all()
    payments = db.query(Payment).filter(Payment.status == PaymentStatus.confirmed).all()

    wb = openpyxl.Workbook()

    # ── Sheet 1: Tenants ──
    ws1 = wb.active
    ws1.title = "Tenant Ledger"
    ws1.sheet_view.showGridLines = False
    for col, width in zip(['A','B','C','D','E','F'], [28,20,12,18,16,16]):
        ws1.column_dimensions[col].width = width

    ws1.merge_cells('A1:F1')
    ws1['A1'] = "Full Portfolio Report — Tenant Ledger"
    ws1['A1'].font      = Font(bold=True, size=11, color=EXCEL_NAVY)
    ws1['A1'].alignment = Alignment(horizontal='center')

    row = 3
    for c, h in enumerate(['Name','Phone','Unit','Balance','Lease End','Status'], 1):
        apply_style(ws1.cell(row, c, h), excel_header_style())
    ws1.row_dimensions[row].height = 18
    row += 1

    for i, t in enumerate(tenants):
        unit = db.query(Unit).filter(Unit.id == t.unit_id).first()
        data = [
            t.name, t.phone_number,
            unit.unit_number if unit else '—',
            f"KES {float(t.balance or 0):,.0f}",
            str(t.lease_end_date),
            'Cleared' if t.balance <= 0 else 'Outstanding'
        ]
        style = excel_row_style(i)
        for c, v in enumerate(data, 1):
            apply_style(ws1.cell(row, c, v), style)
        ws1.row_dimensions[row].height = 15
        row += 1

    # ── Sheet 2: Units ──
    ws2 = wb.create_sheet("Units Overview")
    ws2.sheet_view.showGridLines = False
    for col, width in zip(['A','B','C','D'], [16,20,16,16]):
        ws2.column_dimensions[col].width = width

    ws2.merge_cells('A1:D1')
    ws2['A1'] = "Units Overview"
    ws2['A1'].font      = Font(bold=True, size=11, color=EXCEL_NAVY)
    ws2['A1'].alignment = Alignment(horizontal='center')

    row = 3
    for c, h in enumerate(['Unit No.','Monthly Rent','Status','Occupancy'], 1):
        apply_style(ws2.cell(row, c, h), excel_header_style())
    ws2.row_dimensions[row].height = 18
    row += 1

    active_tenant_unit_ids = {t.unit_id for t in tenants}
    for i, u in enumerate(units):
        data = [
            u.unit_number,
            f"KES {float(u.rent_amount or 0):,.0f}",
            str(u.status.value) if hasattr(u.status,'value') else str(u.status),
            'Occupied' if u.id in active_tenant_unit_ids else 'Vacant'
        ]
        style = excel_row_style(i)
        for c, v in enumerate(data, 1):
            apply_style(ws2.cell(row, c, v), style)
        ws2.row_dimensions[row].height = 15
        row += 1

    # ── Sheet 3: Payments ──
    ws3 = wb.create_sheet("All Payments")
    ws3.sheet_view.showGridLines = False
    for col, width in zip(['A','B','C','D','E','F'], [24,14,20,24,20,20]):
        ws3.column_dimensions[col].width = width

    ws3.merge_cells('A1:F1')
    ws3['A1'] = "All Confirmed Payments"
    ws3['A1'].font      = Font(bold=True, size=11, color=EXCEL_NAVY)
    ws3['A1'].alignment = Alignment(horizontal='center')

    row = 3
    for c, h in enumerate(['Receipt No.','Tenant ID','Amount','Method','Phone','Date'], 1):
        apply_style(ws3.cell(row, c, h), excel_header_style())
    ws3.row_dimensions[row].height = 18
    row += 1

    for i, p in enumerate(sorted(payments, key=lambda x: x.timestamp or datetime.min)):
        data = [
            p.mpesa_receipt_number or '—',
            f"#{p.tenant_id}",
            f"KES {float(p.amount or 0):,.0f}",
            p.payment_method.value if hasattr(p.payment_method,'value') else str(p.payment_method),
            p.phone_number or '—',
            p.timestamp.strftime('%d %b %Y %H:%M') if p.timestamp else '—'
        ]
        style = excel_row_style(i)
        for c, v in enumerate(data, 1):
            apply_style(ws3.cell(row, c, v), style)
        ws3.row_dimensions[row].height = 15
        row += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer